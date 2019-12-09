#模拟生成随机成绩数据

import openpyxl
from openpyxl import Workbook
import random

'''假设某学校所有课程每学期允许多次考试，学生随时可以参加考试，系统自动将每次成绩添加到Excel文件（包含3列：姓名
、课程、成绩）中，现期末要求统计所有学生每门课程的最高成绩'''

#生成随机数据
def generateRandomInformation(filename):
    workbook=Workbook()
    worksheet=workbook.worksheets[0]
    worksheet.append(['姓名','课程','成绩'])
    #中文名字中的第一、第二、第三个字
    first=tuple('魏吴张曹')
    middle=tuple('巍伯权操')
    last=tuple('权帅牛')
    #课程名名字
    subjects=('语文','数学','英语')
    #随机生成200个数据
    for i in range(200):
        line=[]
        r=random.randint(1,100)
        name=random.choice(first)
        #按一定概率生成只有两个字的中文名字
        if r>50:
            name=name+random.choice(middle)
        name=name+random.choice(last)
        #依次生成姓名、课程名名字和成绩
        line.append(name)
        line.append(random.choice(subjects))
        line.append(random.randint(0,100))
        worksheet.append(line)
    #保存数据，生成Excel2007格式的文件
    workbook.save(filename)

def getResult(oldfile,newfile):
    #用于存放结果数据的字典
    result=dict()
    #打开原始数据
    workbook=openpyxl.load_workbook(oldfile)
    worksheet=workbook.worksheets[0]
    #遍历原始数据
    for row in list(worksheet.rows)[1:]:
        #姓名，课程名称，本次成绩
        name,subject,grade=row[0].value,row[1].value,row[2].value
        #获取当前姓名对应的课程名称和成绩信息
        #如果result字典中不包含，则返回空字典
        t=result.get(name,{})
        #获取当前学生当前课程的成绩，若不存在，返回0
        f=t.get(subject,0)
        #只保学生课程的最高成绩
        if grade>f:
            t[subject]=grade
            result[name]=t
        #创建Excel文件
        workbook1=Workbook()
        worksheet1=workbook1.worksheets[0]
        worksheet1.append(['姓名','课程','成绩'])
        #将result字典中的结果数据写入Excel文件
        for name,t in result.items():
            for subject,grade in t.items():
                worksheet1.append([name,subject,grade])
        workbook1.save(newfile)

if __name__=='__main__':
    oldfile='student.xlsx'
    newfile='result.xlsx'
    generateRandomInformation(oldfile)
    getResult(oldfile,newfile)
    





















        
























        
