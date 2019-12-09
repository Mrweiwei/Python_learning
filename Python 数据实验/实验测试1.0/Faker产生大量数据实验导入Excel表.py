from faker import Faker
from faker.providers import BaseProvider
import random
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

fake = Faker('zh_CN')
class MyProvider(BaseProvider):
    def goods(self):
        goods=['防水迷你夜视接手机望远镜','智能定时开关插排','智能手环手表','家用无线门铃','智能蓝牙偏光眼镜','迷你手机小话筒','迷你金属小钢炮','智能体重秤','智能蓝牙防丢器','智能人体感应小夜灯']
        return random.choice(goods)
    def position(self):
        position=['百度糯米','百度外卖','必胜客','当当网','达美乐比萨','大众点评','饿了么','1号店','凡客诚品','国美电器','京东','聚美优品','酒仙网','吉野家','肯德基','麦当劳','驴妈妈旅游网','蘑菇街','美团网','美团外卖','苏宁易购','淘宝网','唯品会','网易考拉海购','亚马逊','我买网']
        return random.choice(position)
fake.add_provider(MyProvider)

def generateInformation(filename):
    workbook=Workbook()
    worksheet=workbook.worksheets[0]
    worksheet.append(['时间','地点','商品名称'])    

    for i in range(1,100000):
        line=[]
        line.append(fake.date_this_month())
        line.append(fake.position())
        line.append(fake.goods())
        worksheet.append(line)
    #保存数据，生成Excel2007格式的文件
    workbook.save(filename)
        
if __name__=='__main__':
    oldfile='仿订单实验数据.xlsx'
    generateInformation(oldfile)
    wb = load_workbook('仿订单实验数据.xlsx') 
    ws = wb[wb.sheetnames[0]]
    # 调整列宽
    ws.column_dimensions['A'].width = 15.0
    ws.column_dimensions['B'].width = 25.0
    ws.column_dimensions['C'].width = 35.0
    wb.save('仿订单实验数据.xlsx')

