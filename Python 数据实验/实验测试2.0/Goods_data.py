#操作2007年之后版本用的Excel表格库
import openpyxl
from openpyxl import load_workbook


'''第一部分搭建5万多条数据的数据库'''
#读取数据
wb=load_workbook("Database.xlsx")
ws=wb.worksheets[0]

#定义一个商品列表Goods_list，这是一个字典列表，列表中的每个元素都是一个字典
Goods_list=[]

'''a={"title":"asd"}
Goods_list.append(a)
print(Goods_list[0]["title"])'''

#商品名称
Shopping_name= []
for _ in ws["A"]:
  Shopping_name.append(_.value)
#print(Shopping_name[0])
#商品金额
Shopping_money=[]
for _ in ws["C"]:
    Shopping_money.append(_.value)
#print(Shopping_money)

#详情地址
Shopping_url=[]
for _ in ws["B"]:
    Shopping_url.append(_.value)
#print(Shopping_url)

#所属类别
Category_id=[]
for _ in ws["D"]:
    Category_id.append(_.value)
#来源
Source_id=[]
for _ in ws["E"]:
    Source_id.append(_.value)

#向字典列表中插入一个个字典     
for m in range(len(Shopping_name)):
    a={}
    a["title"]=Shopping_name[m]
    a["price"]=Shopping_money[m]
    a["url"]=Shopping_url[m]
    a["category_id"]=Category_id[m]
    a["source_id"]=Source_id[m]
    Goods_list.append(a)

#print(Goods_list[51043])
    
'''第二部分对数据库中的数据按照对应的类别ID筛选分类到各自类别的列表中'''
#声明分类列表
家用电器=[]
手机=[]
数码=[]
电脑周边=[]
办公=[]
家具=[]
家居=[]
家装=[]
运营商=[]
厨具=[]
男装=[]
女装=[]
童装=[]
内衣=[]
美妆个护=[]
宠物=[]
女鞋=[]
箱包=[]
钟表=[]
珠宝首饰=[]
运动服饰=[]
男鞋=[]
户外=[]
汽车=[]
汽车用品=[]
母婴=[]
玩具乐器=[]
酒类=[]
生鲜=[]
特产=[]
食物饮品=[]
礼品鲜花=[]
农资绿植=[]
医疗保健=[]
计生情趣=[]
图书=[]
音像=[]
机票=[]
电影票=[]
旅游=[]
酒店=[]
理财=[]
外卖=[]
其它=[]
网游=[]
保险=[]

#将构造的数据库中的数据按照其对应的category_id进行筛选分类
for _ in (Goods_list):
    if _["category_id"]==1:
        手机.append(_)
#print(家用电器[34]["title"])
    elif _["category_id"]==0:
        家用电器.append(_)
    elif _["category_id"]==3:
        数码.append(_)
    elif _["category_id"]==4:
        电脑周边.append(_)    
    elif _["category_id"]==5:
        办公.append(_)        
    elif _["category_id"]==7:
        家具.append(_)
    elif _["category_id"]==6:
        家居.append(_)
    elif _["category_id"]==8:
        家装.append(_)
    elif _["category_id"]==2:
        运营商.append(_)
    elif _["category_id"]==9:
        厨具.append(_)    
    elif _["category_id"]==10:
        男装.append(_)    
    elif _["category_id"]==11:
        女装.append(_)       
    elif _["category_id"]==12:
        童装.append(_)
    elif _["category_id"]==13:
        内衣.append(_)
    elif _["category_id"]==14:
        美妆个护.append(_)
    elif _["category_id"]==15:
        宠物.append(_)
    elif _["category_id"]==16:
        女鞋.append(_)
    elif _["category_id"]==17:
        箱包.append(_)
    elif _["category_id"]==18:
        钟表.append(_)
    elif _["category_id"]==19:
        珠宝首饰.append(_)
    elif _["category_id"]==21:
        运动服饰.append(_)
    elif _["category_id"]==20:
        男鞋.append(_)
    elif _["category_id"]==22:
        户外.append(_)
    elif _["category_id"]==23:
        汽车.append(_)
    elif _["category_id"]==24:
        汽车用品.append(_)
    elif _["category_id"]==25:
        母婴.append(_)
    elif _["category_id"]==26:
        玩具乐器.append(_)
    elif _["category_id"]==28:
        酒类.append(_)
    elif _["category_id"]==29:
        生鲜.append(_)
    elif _["category_id"]==30:
        特产.append(_)
    elif _["category_id"]==27:
        食物饮品.append(_)
    elif _["category_id"]==31:
        礼品鲜花.append(_)
    elif _["category_id"]==32:
        农资绿植.append(_)
    elif _["category_id"]==33:
        医疗保健.append(_)
    elif _["category_id"]==34:
        计生情趣.append(_)
    elif _["category_id"]==35:
        图书.append(_)
    elif _["category_id"]==36:
        音像.append(_)
    elif _["category_id"]==38:
        机票.append(_)
    elif _["category_id"]==37:
        电影票.append(_)
    elif _["category_id"]==40:
        旅游.append(_)
    elif _["category_id"]==46:
        其它.append(_)
    elif _["category_id"]==42:
        理财.append(_)
    elif _["category_id"]==46:
        其它.append(_)
    elif _["category_id"]==44:
        网游.append(_)
    elif _["category_id"]==46:
        其它.append(_)
    elif _["category_id"]==45:
        保险.append(_)
    elif _["category_id"]==47:
        外卖.append(_)
    elif _["category_id"]==40:
        旅游.append(_)
    elif _["category_id"]==39:
        酒店.append(_)
    elif _["category_id"]==40:
        旅游.append(_)



'''
#随机拿10条外卖的数据
for i in range(10):
    print(外卖[i]["title"])
#随机拿20条旅游数据
for i in range(20):
    print(旅游[i]["title"])'''

'''第三部分将45个小类合并成5个大类用、衣、食、玩和其它'''
玩=[]
玩=电影票+机票+酒店+旅游

用=[]
用=家用电器+手机+运营商+数码+电脑周边+办公+家居+家具+家装+厨具+钟表+珠宝首饰+汽车+汽车用品+母婴+玩具乐器+计生情趣+图书+音像

衣=[]
衣=男装+女装+童装+内衣+美妆个护+女鞋+箱包+男鞋+运动服饰+户外

食=[]
食=食物饮品+酒类+生鲜+特产+医疗保健+外卖

其它=[]
其它=宠物+礼品鲜花+农资绿植+理财+保险+其它






        




        









    
        
        

    
