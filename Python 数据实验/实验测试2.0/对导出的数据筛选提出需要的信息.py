import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import NamedStyle, Font, Border, Side,PatternFill
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import pandas as pd
import numpy as np
from openpyxl.styles import Font
import Goods_data

Goods_list=Goods_data.Goods_list
#筛选获得的数据进行读取
wb=load_workbook("Mysql数据库中导出的数据.xlsx")
ws=wb.worksheets[0]

#商品名称
Shopping_name= []
for i in ws["C"]:
  Shopping_name.append(i.value)
#print(list_sheet1_column_C )

#商品数量
Shopping_num=[]
for _ in ws["D"]:
    Shopping_num.append(_.value)
#print(Shopping_num)      

#商品金额
Shopping_money=[]
for _ in ws["E"]:
    Shopping_money.append(_.value)
#print(Shopping_money)

#利用pandas库建立dataframe
learn_data={"Name":[],"Num":[],"Price":[]}
for i in range(1,len(Shopping_name)):
  learn_data["Name"].append(Shopping_name[i])
  learn_data["Num"].append(Shopping_num[i])
  learn_data["Price"].append(Shopping_money[i])
  
Goods=pd.DataFrame(learn_data)
#按照名称分组
Order=Goods.groupby(["Name","Price"])["Num"].sum()


  
'''Shopping_list=[]
for i in range(len(Goods_list)):
  Shopping_list.append(Goods_list[i]["title"])
  
if Order.index[0] in Shopping_list:
  print("yes")
else:
  print("no")'''





Order.to_excel("处理数据.xlsx")
wb1= load_workbook('处理数据.xlsx') 
ws1= wb1[wb1.sheetnames[0]]
ws1["C1"]="Num"
ws1["D1"]="Total"

#设置单元格格式
ft = Font(name=u'微软雅黑',
    size=11,
    bold=False,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='FF000000')

alignment=Alignment(horizontal='general',
        vertical='bottom',
        text_rotation=0,
        wrap_text=False,
        shrink_to_fit=False,
        indent=0)

ft1= Font(name=u'微软雅黑',
    size=11,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='FF000000')

ws1["D1"].font=ft1

for row in range(2,len(Order.index)+2):
    row=str(row)
    ws1["D"+row]='=C{0}*B{0}'.format(row)
    ws1["A"+row].font = ft
    ws1["A"+row].alignment =alignment
    ws1["B"+row].font = ft
    ws1["B"+row].alignment =alignment
ws1.column_dimensions['A'].width = 100.0
ws1.column_dimensions['B'].width = 10.0
wb1.save('处理数据.xlsx')





'''
new_data = pd.DataFrame(columns=["Name", "Num"])
for key,value in Order:
  new_data = pd.concat([new_data, value])
#print(new_data)
goods=new_data.pivot_table('Num',columns=['Name'],aggfunc=sum)
#print(a)
'''









'''
#工作表2
ws1=wb.worksheets[1]
#print(ws1["A"][0].value)
#ws1["A"+str(1)].value=Shopping_name[5]
#存入需要的数据
for i in range(1,len(Order["name"])+1):
    ws1["A"+str(i)].value=Shopping_name[i-1]
ws1.column_dimensions['A'].width = 70.0

for i in range(1,len(Shopping_money)+1):
    ws1["B"+str(i)].value=Shopping_money[i-1]
ws1.column_dimensions['B'].width = 10.0

for i in range(1,len(Shopping_num)+1):
    ws1["C"+str(i)].value=Shopping_num[i-1]
ws1.column_dimensions['C'].width = 10.0
'''

'''
#追加一列统计总金额
ws1["D1"]="总金额"
for row in range(2,len(Shopping_num)+1):
    row=str(row)
    ws1["D"+row]='=C{0}*B{0}'.format(row)
wb.save("Mysql数据库中导出的数据.xlsx")'''
    
    



    
    
















    
