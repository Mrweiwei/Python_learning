'''                       邮箱库                    '''
#操作2007年之后版本用的Excel表格库
import openpyxl
from openpyxl import load_workbook
import random
from openpyxl import Workbook
#创件一个Excel文件
workbook=Workbook()
#创建一个工作表
worksheet=workbook.worksheets[0]

'''第一部分搭建七千条仿邮件的数据库'''
#读取数据
wb=load_workbook("email.xlsx")
ws=wb.worksheets[0]

#定义一个邮件列表，是一个字典列表
Mail_list=[]

#邮件主题名称
Mail_theme= []
for _ in ws["A"]:
  Mail_theme.append(_.value)

Mail_category_id=[]
for _ in ws["B"]:
  Mail_category_id.append(_.value)

#向字典列表中插入一个个字典
for m in range(len(Mail_theme)):
  a={}
  a["theme"]=Mail_theme[m]
  a["category_id"]=Mail_category_id[m]
  Mail_list.append(a)


#測試：
#for i in range(20):
#  print(Mail_list[i])

#根据category_id分别装进正常邮件和垃圾邮件库
正常邮箱=[]
垃圾邮箱=[]
for _ in Mail_list:
    if _["category_id"]==2:
        垃圾邮箱.append(_)
    elif _["category_id"]==0:
        正常邮箱.append(_)
#测试:
#for i in range(10):
#    print(垃圾邮箱[i])
#    print(正常邮箱[i])
        












"""
#or '【' in Mail_list[i]["theme"] or'广告' in Mail_list[i]["theme"] or'#' in Mail_list[i]["theme"] or'*' in Mail_list[i]["theme"]

#简单分类
for i in range(1000):
  #print(Mail_list[i]["theme"])
  if '[' in Mail_list[i]["theme"]or '【' in Mail_list[i]["theme"] or'广告' in Mail_list[i]["theme"] or'#' in Mail_list[i]["theme"] or'*' in Mail_list[i]["theme"]:
    Mail_list[i]["category_id"]=2
  else:
    Mail_list[i]["category_id"]=0
  line=[]
  line.append(Mail_list[i]["theme"])
  line.append(Mail_list[i]["category_id"])
  worksheet.append(line)

#保存到excel表格中
workbook.save("emaila.xlsx")"""
