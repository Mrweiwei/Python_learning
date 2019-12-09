import pymysql
#引入操作Excel表格的openpyxl库
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill

#创件一个Excel文件
workbook=Workbook()
#创建一个工作表
worksheet=workbook.worksheets[0]
#表的第一行
worksheet.append(['ID','订单号','商品名称','数量','金额','支付方式','购买时间','商品网址','购买渠道'])  
# 打开数据库连接
conn = pymysql.connect("localhost","root","","runoob_db" )
# 使用cursor()方法获取操作游标 
cursor =conn.cursor()
# SQL 查询语句
sql = "SELECT * FROM GOODS_DATA"
try:
    #执行SQL语句
    cursor.execute(sql)
    #获取所有记录列表
    results=cursor.fetchall()
    for row in results:
        line=[]
        '''
        id=row[0]
        order_id=row[1]
        goods=row[2]
        num=row[3]
        payway=row[4]
        time=row[5]
        link_href=row[6]
        position=row[7]
        #打印地址检测
        #print("position:%s"%(position))'''
        line.append(row[0])
        line.append(row[1])
        line.append(row[2])
        line.append(row[3])
        line.append(row[4])
        line.append(row[5])
        line.append(row[6])
        line.append(row[7])
        line.append(row[8])
        worksheet.append(line)
except:
    print ("Error: unable to fetch data")
   
# 关闭数据库连接
conn.close()
#保存数据，生成Excel2007格式的文件
workbook.save("Mysql数据库中导出的数据.xlsx")
wb = load_workbook('Mysql数据库中导出的数据.xlsx') 
ws = wb[wb.sheetnames[0]]
# 调整列宽
ws.column_dimensions['A'].width = 10.0
ws.column_dimensions['B'].width = 20.0
ws.column_dimensions['C'].width = 70.0
ws.column_dimensions['D'].width = 10.0
ws.column_dimensions['E'].width = 10.0
ws.column_dimensions['F'].width = 15.0
ws.column_dimensions['G'].width = 15.0
ws.column_dimensions['H'].width = 100.0
ws.column_dimensions['I'].width = 15.0
wb.save('Mysql数据库中导出的数据.xlsx')
print("数据导出成功！")













        
